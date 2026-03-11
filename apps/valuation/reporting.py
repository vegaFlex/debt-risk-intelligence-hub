import io

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas


def build_valuation_report_summary(portfolio, preview, latest_valuation=None, comparison_rows=None):
    return {
        'portfolio': portfolio,
        'preview': preview,
        'latest_valuation': latest_valuation,
        'comparison_rows': comparison_rows or [],
        'top_factors': preview['factors'][:6],
        'top_scenarios': preview['scenarios'],
    }


def build_valuation_excel_report(summary):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Valuation Summary'

    portfolio = summary['portfolio']
    preview = summary['preview']

    ws.append(['Portfolio', portfolio.name])
    ws.append(['Source Company', portfolio.source_company])
    ws.append(['Currency', portfolio.currency])
    ws.append(['Face Value', float(portfolio.face_value)])
    ws.append(['Expected Recovery %', float(preview['expected_recovery_rate'])])
    ws.append(['Expected Collections', float(preview['expected_collections'])])
    ws.append(['Recommended Bid %', float(preview['recommended_bid_pct'])])
    ws.append(['Recommended Bid Amount', float(preview['recommended_bid_amount'])])
    ws.append(['Projected ROI %', float(preview['projected_roi'])])
    ws.append(['Confidence Score', float(preview['confidence_score'])])
    ws.append([])

    ws.append(['Top Factors'])
    ws.append(['Factor', 'Weight', 'Value', 'Explanation'])
    for factor in summary['top_factors']:
        ws.append([factor['factor_name'], float(factor['factor_weight']), factor['factor_value'], factor['explanation']])

    ws_scenarios = wb.create_sheet(title='Scenario Analysis')
    ws_scenarios.append(['Bid %', 'Bid Amount', 'Expected Profit', 'ROI %', 'Break-Even Recovery %', 'Recommended'])
    for scenario in summary['top_scenarios']:
        ws_scenarios.append([
            float(scenario['bid_pct']),
            float(scenario['bid_amount']),
            float(scenario['expected_profit']),
            float(scenario['roi']),
            float(scenario['break_even_recovery']),
            'Yes' if scenario['is_recommended'] else 'No',
        ])

    if summary['comparison_rows']:
        ws_history = wb.create_sheet(title='Saved Runs')
        ws_history.append(['Saved At', 'Method', 'Expected Recovery %', 'Recommended Bid %', 'Projected ROI %', 'Confidence'])
        for row in summary['comparison_rows']:
            valuation = row['valuation']
            ws_history.append([
                valuation.created_at.strftime('%Y-%m-%d %H:%M'),
                valuation.get_valuation_method_display(),
                float(valuation.expected_recovery_rate),
                float(valuation.recommended_bid_pct),
                float(valuation.projected_roi),
                float(valuation.confidence_score),
            ])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def build_valuation_pdf_report(summary):
    portfolio = summary['portfolio']
    preview = summary['preview']

    output = io.BytesIO()
    pdf = canvas.Canvas(output, pagesize=A4)
    width, height = A4
    y = height - 40

    pdf.setFont('Helvetica-Bold', 14)
    pdf.drawString(40, y, 'Debt Portfolio Valuation Memo')
    y -= 20
    pdf.setFont('Helvetica', 10)
    pdf.drawString(40, y, f'Portfolio: {portfolio.name}')
    y -= 14
    pdf.drawString(40, y, f'Source: {portfolio.source_company or "N/A"} | Currency: {portfolio.currency}')
    y -= 22

    pdf.setFont('Helvetica-Bold', 11)
    pdf.drawString(40, y, 'Core Recommendation')
    y -= 16
    pdf.setFont('Helvetica', 10)
    lines = [
        f'Expected Recovery: {preview["expected_recovery_rate"]}%',
        f'Expected Collections: {preview["expected_collections"]}',
        f'Recommended Bid: {preview["recommended_bid_pct"]}% ({preview["recommended_bid_amount"]})',
        f'Projected ROI: {preview["projected_roi"]}%',
        f'Confidence: {preview["confidence_score"]}',
    ]
    for line in lines:
        pdf.drawString(50, y, line)
        y -= 14

    y -= 8
    pdf.setFont('Helvetica-Bold', 11)
    pdf.drawString(40, y, 'Key Drivers')
    y -= 16
    pdf.setFont('Helvetica', 9)
    for factor in summary['top_factors']:
        pdf.drawString(45, y, f"{factor['factor_name']}: {factor['factor_value']} ({factor['factor_weight']})")
        y -= 12
        if y < 60:
            pdf.showPage()
            y = height - 40
            pdf.setFont('Helvetica', 9)

    y -= 8
    pdf.setFont('Helvetica-Bold', 11)
    pdf.drawString(40, y, 'Scenario Analysis')
    y -= 16
    pdf.setFont('Helvetica', 9)
    for scenario in summary['top_scenarios']:
        label = 'Recommended' if scenario['is_recommended'] else 'Alternative'
        pdf.drawString(45, y, f"{scenario['bid_pct']}% | ROI {scenario['roi']}% | Break-even {scenario['break_even_recovery']}% | {label}")
        y -= 12
        if y < 60:
            pdf.showPage()
            y = height - 40
            pdf.setFont('Helvetica', 9)

    pdf.save()
    return output.getvalue()
